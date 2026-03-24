#!/usr/bin/env python3
"""
从 .xlsx 中按表头列名导出内嵌图片（仅导出锚点落在该列上的图）。

原理：解析 xl/drawings/*.xml 里 twoCellAnchor/oneCellAnchor 的 xdr:from/xdr:col
（0 起算：0=A, 1=B），与表头「画面」等列对齐后，从压缩包复制 xl/media/*。

依赖：openpyxl（仅用于读表头定位列号）

用法:
  python3 export_xlsx_column_images.py "/path/to/脚本.xlsx"
  python3 export_xlsx_column_images.py 脚本范例.xlsx -o ~/Desktop/画面列配图 --column 画面
  python3 export_xlsx_column_images.py 脚本范例.xlsx --sheet "副本写作_5" --header-row 2
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _col_letter_to_index(col: str) -> int:
    col = col.upper().strip()
    n = 0
    for c in col:
        if not ("A" <= c <= "Z"):
            break
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _index_to_col_letter(idx: int) -> str:
    s = ""
    idx += 1
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _load_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    raw = zf.read(rels_path)
    root = ET.fromstring(raw)
    out: dict[str, str] = {}
    for rel in root.findall("pr:Relationship", NS):
        rid = rel.get("Id")
        tgt = rel.get("Target")
        if rid and tgt:
            out[rid] = tgt.replace("\\", "/")
    return out


def _resolve_part_path(base_xml: str, target: str) -> str:
    p = os.path.normpath(os.path.join(os.path.dirname(base_xml), target))
    return p.replace("\\", "/")


def _workbook_sheet_path(zf: zipfile.ZipFile, sheet_name: str | None, sheet_index: int) -> tuple[str, str]:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    wb_rels = _load_rels(zf, "xl/_rels/workbook.xml.rels")
    sheets = wb_root.findall("m:sheets/m:sheet", NS)
    if not sheets:
        raise ValueError("workbook 中未找到 sheet")

    if sheet_name:
        chosen = None
        for sh in sheets:
            if sh.get("name") == sheet_name:
                chosen = sh
                break
        if chosen is None:
            names = [sh.get("name") for sh in sheets]
            raise ValueError(f"未找到工作表「{sheet_name}」。已有: {names}")
    else:
        if sheet_index < 0 or sheet_index >= len(sheets):
            raise ValueError(f"sheet-index 越界: {sheet_index}（共 {len(sheets)} 张）")
        chosen = sheets[sheet_index]

    rid = chosen.get(f"{{{NS['r']}}}id")
    if not rid or rid not in wb_rels:
        raise ValueError("无法解析工作表关系")
    path = _resolve_part_path("xl/workbook.xml", wb_rels[rid])
    return path, chosen.get("name") or path


def find_column_index(xlsx_path: Path, ws_title: str, header_row: int, column_name: str) -> int:
    try:
        import openpyxl
    except ImportError:
        print("需要 openpyxl：python3 -m pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if ws_title not in wb.sheetnames:
            raise ValueError(f"工作表「{ws_title}」不存在，可选: {wb.sheetnames}")
        ws = wb[ws_title]
        for cell in ws[header_row]:
            val = cell.value
            if val is None:
                continue
            if str(val).strip() == column_name.strip():
                return _col_letter_to_index(cell.column_letter)
    finally:
        wb.close()
    raise ValueError(f"在第 {header_row} 行未找到表头列「{column_name}」")


def _drawing_target_from_sheet(zf: zipfile.ZipFile, sheet_xml: str) -> str | None:
    root = ET.fromstring(zf.read(sheet_xml))
    for el in root.iter():
        if el.tag.endswith("}drawing"):
            rid = el.get(f"{{{NS['r']}}}id")
            if not rid:
                continue
            rels_path = sheet_xml.replace("worksheets/", "worksheets/_rels/").replace(".xml", ".xml.rels")
            rels = _load_rels(zf, rels_path)
            if rid not in rels:
                continue
            return _resolve_part_path(sheet_xml, rels[rid])
    return None


def _iter_anchors_with_embed(root: ET.Element):
    xdr = NS["xdr"]
    for tag in ("twoCellAnchor", "oneCellAnchor"):
        for anchor in root.findall(f".//{{{xdr}}}{tag}"):
            frm = anchor.find(f"{{{xdr}}}from")
            if frm is None:
                continue
            col_el = frm.find(f"{{{xdr}}}col")
            row_el = frm.find(f"{{{xdr}}}row")
            if col_el is None or row_el is None or col_el.text is None or row_el.text is None:
                continue
            col = int(col_el.text)
            row = int(row_el.text)
            pic = anchor.find(f"{{{xdr}}}pic")
            if pic is None:
                continue
            nv = pic.find(f"{{{xdr}}}nvPicPr")
            pic_name = ""
            if nv is not None:
                cnv = nv.find(f"{{{xdr}}}cNvPr")
                if cnv is not None:
                    pic_name = cnv.get("name") or ""
            blip = pic.find(f".//{{{NS['a']}}}blip")
            if blip is None:
                continue
            embed = blip.get(f"{{{NS['r']}}}embed")
            if not embed:
                continue
            yield row, col, embed, pic_name


def export_column_images(
    xlsx_path: Path,
    out_dir: Path,
    column_name: str,
    header_row: int,
    sheet_name: str | None,
    sheet_index: int,
) -> dict:
    xlsx_path = xlsx_path.expanduser().resolve()
    out_dir = out_dir.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        sheet_xml, display_name = _workbook_sheet_path(zf, sheet_name, sheet_index)
        target_col = find_column_index(xlsx_path, display_name, header_row, column_name)

        drawing_xml = _drawing_target_from_sheet(zf, sheet_xml)
        items: list[dict] = []
        if not drawing_xml:
            manifest = {
                "source_xlsx": str(xlsx_path),
                "sheet": display_name,
                "filter_column": column_name,
                "filter_col_index_0based": target_col,
                "filter_col_letter": _index_to_col_letter(target_col),
                "header_row": header_row,
                "count": 0,
                "images": [],
                "note": "该工作表没有 drawing 节点",
            }
            (out_dir / "manifest.json").write_text(
                json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            return manifest

        droot = ET.fromstring(zf.read(drawing_xml))
        drels_path = drawing_xml.replace("drawings/", "drawings/_rels/").replace(".xml", ".xml.rels")
        drels = _load_rels(zf, drels_path)

        for row0, col0, embed_rid, pic_name in _iter_anchors_with_embed(droot):
            if col0 != target_col:
                continue
            if embed_rid not in drels:
                continue
            media_path = _resolve_part_path(drawing_xml, drels[embed_rid])
            if media_path not in zf.namelist():
                continue
            excel_row = row0 + 1
            ext = Path(media_path).suffix or ".bin"
            safe = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", pic_name)[:50] if pic_name else ""
            stem = safe or Path(media_path).stem
            out_name = f"row{excel_row:03d}_{_index_to_col_letter(target_col)}_{stem}{ext}"
            out_path = out_dir / out_name
            with zf.open(media_path) as src, open(out_path, "wb") as dst:
                shutil.copyfileobj(src, dst)
            items.append(
                {
                    "excel_row": excel_row,
                    "column": _index_to_col_letter(target_col),
                    "column_header": column_name,
                    "embedded": media_path,
                    "saved_as": out_name,
                    "pic_name_excel": pic_name,
                }
            )

        items.sort(key=lambda x: (x["excel_row"], x["saved_as"]))
        manifest = {
            "source_xlsx": str(xlsx_path),
            "sheet": display_name,
            "filter_column": column_name,
            "filter_col_index_0based": target_col,
            "filter_col_letter": _index_to_col_letter(target_col),
            "header_row": header_row,
            "drawing_xml": drawing_xml,
            "count": len(items),
            "images": items,
        }
        (out_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return manifest


def main() -> None:
    ap = argparse.ArgumentParser(description="导出 xlsx 指定表头列上的内嵌图片（按锚点列过滤）")
    ap.add_argument("xlsx", type=Path, help="输入 .xlsx")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出目录（默认：<xlsx同目录>/<stem>_列画面配图）",
    )
    ap.add_argument("--column", default="画面", help="表头列名，默认「画面」")
    ap.add_argument("--header-row", type=int, default=2, help="表头所在行（1-based），默认 2")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--sheet", default=None, help="工作表名称")
    g.add_argument("--sheet-index", type=int, default=None, help="工作表序号（0 起）；不设则默认第一张")
    args = ap.parse_args()

    out = args.output
    if out is None:
        col_short = re.sub(r'[\\/:*?"<>|]', "_", args.column)[:20]
        out = args.xlsx.parent / f"{args.xlsx.stem}_列{col_short}配图"

    sheet_name = args.sheet
    sheet_index = args.sheet_index if args.sheet_index is not None else 0

    try:
        man = export_column_images(
            args.xlsx,
            out,
            args.column,
            args.header_row,
            sheet_name,
            sheet_index,
        )
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"工作表: {man['sheet']}")
    print(f"列「{man['filter_column']}」→ {_index_to_col_letter(man['filter_col_index_0based'])} 列（0-based={man['filter_col_index_0based']}）")
    print(f"共导出 {man['count']} 张 → {out.resolve()}")
    print(f"清单: {out / 'manifest.json'}")
    for it in man["images"][:20]:
        print(f"  Excel行{it['excel_row']}: {it['saved_as']}")
    if man["count"] > 20:
        print(f"  ... 共 {man['count']} 张")


if __name__ == "__main__":
    main()
