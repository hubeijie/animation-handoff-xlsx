#!/usr/bin/env python3
"""从 JSON 生成「美术统筹表」.xlsx，版式对齐「美术统筹表范例.xlsx」（三栏并列），支持配图嵌入。"""

from __future__ import annotations

import argparse
import json
import sys
from copy import copy
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖：python3 -m pip install openpyxl pillow", file=sys.stderr)
    sys.exit(1)

# 无「分镜截图」列；保留「脚本截图」「参考图」。
# 场景列 1–6，与道具列之间**仅隔 1 列空白**（第 7 列）；道具 8–12 与角色 14–18 之间仅隔第 13 列。
SCENE_COLS = {
    "序号": 1,
    "背景文件名": 2,
    "场景名称": 3,
    "场景描述": 4,
    "脚本截图": 5,
    "参考图": 6,
}
PROP_COLS = {
    "序号": 8,
    "道具名称": 9,
    "道具描述": 10,
    "脚本截图": 11,
    "参考图": 12,
}
CHAR_COLS = {
    "序号": 14,
    "角色名称": 15,
    "角色描述": 16,
    "脚本截图": 17,
    "参考图": 18,
}

IMAGE_COL_KEYS = ("脚本截图", "参考图")

# 明细行中整格仅为占位符时写入空单元格（不写「—」等）
_PLACEHOLDER_STRIPS = frozenset(
    {"—", "–", "-", "－", "―", "n/a", "N/A", "无", "暂无", "无。", "—。"}
)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 嵌图时：宽度不超过所在列（避免盖住右侧「参考图」等列），高度上限避免单行无限拉高。
# Excel 列宽单位为「默认字体下字符数」，换算像素为近似值（Calibri 11 下常用经验系数）。
_EXCEL_WIDTH_UNITS_TO_PX = 7.0
_IMAGE_H_PADDING_PX = 6
_MAX_EMBED_IMAGE_HEIGHT_PX = 260


def _column_max_image_width_px(ws, col_idx: int) -> float:
    """当前列设定宽度下，嵌入图允许的最大像素宽度（留边）。"""
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    units = float(dim.width) if dim is not None and dim.width is not None else 13.0
    px = max(48.0, units * _EXCEL_WIDTH_UNITS_TO_PX - float(_IMAGE_H_PADDING_PX * 2))
    return px


def _px_to_points_h(px: float) -> float:
    """嵌入图高度（像素，按 96dpi）→ Excel 行高（磅）。"""
    return px * 72.0 / 96.0


def _is_reuse_row_name(name: Any) -> bool:
    """名称以 `-复用` 结尾视为复用行，不计入表头数量。"""
    if name is None:
        return False
    s = str(name).strip()
    return s.endswith("-复用")


def _count_new_only_rows(
    scenes: list, props: list, characters: list
) -> tuple[int, int, int]:
    sn = sum(
        1
        for x in scenes
        if isinstance(x, dict) and not _is_reuse_row_name(x.get("场景名称"))
    )
    pn = sum(
        1
        for x in props
        if isinstance(x, dict)
        and "__section__" not in x
        and not _is_reuse_row_name(x.get("道具名称"))
    )
    cn = sum(
        1
        for x in characters
        if isinstance(x, dict) and not _is_reuse_row_name(x.get("角色名称"))
    )
    return sn, pn, cn


def _copy_font(src, dst) -> None:
    if src and src.font:
        dst.font = copy(src.font)


def _style_data_row(ws, row: int, proto_row: int) -> None:
    for c in list(SCENE_COLS.values()) + list(PROP_COLS.values()) + list(CHAR_COLS.values()):
        src = ws.cell(row=proto_row, column=c)
        dst = ws.cell(row=row, column=c)
        _copy_font(src, dst)
        dst.alignment = WRAP_TOP


def _resolve_media_path(raw: Any, base: Path | None) -> Path | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, list):
        raw = raw[0] if raw else None
        if raw is None:
            return None
    if not isinstance(raw, (str, Path)):
        return None
    p = Path(raw).expanduser()
    if not p.is_absolute() and base is not None:
        p = (base / p).resolve()
    else:
        p = p.resolve() if p.is_absolute() else Path.cwd() / p
    return p if p.is_file() else None


def _embed_row_images(
    ws,
    row: int,
    col_map: dict[str, int],
    item: dict,
    base: Path | None,
    max_side: float,
) -> None:
    pics = item.get("配图") or item.get("images") or {}
    if not isinstance(pics, dict):
        return
    max_h_pt: float = 0.0
    for name in IMAGE_COL_KEYS:
        if name not in col_map:
            continue
        raw = pics.get(name)
        if raw is None or raw == "":
            continue
        if isinstance(raw, list) and len(raw) > 1:
            print(
                f"提示：{name} 多张图仅嵌入第一张：{row=} {raw[1:]!r}",
                file=sys.stderr,
            )
        path = _resolve_media_path(raw, base)
        if path is None:
            print(f"警告：找不到配图文件（已跳过）：{raw!r} base={base}", file=sys.stderr)
            continue
        try:
            img = XLImage(str(path))
        except Exception as e:
            print(f"警告：无法读取图片 {path}: {e}", file=sys.stderr)
            continue
        w, h = float(img.width), float(img.height)
        if w <= 0 or h <= 0:
            continue
        col = col_map[name]
        max_w = _column_max_image_width_px(ws, col)
        # 先按列宽限制宽度，再按高度上限与 max_side 限制整体，避免浮动图盖住相邻列文字
        scale = min(max_w / w, float(_MAX_EMBED_IMAGE_HEIGHT_PX) / h, max_side / max(w, h), 1.0)
        img.width = max(1, int(w * scale))
        img.height = max(1, int(h * scale))
        anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img, anchor)
        max_h_pt = max(max_h_pt, _px_to_points_h(float(img.height)))
    if max_h_pt > 0:
        cur = ws.row_dimensions[row].height
        cur_pt = float(cur) if cur is not None else 15.0
        ws.row_dimensions[row].height = max(cur_pt, max_h_pt + 14.0)


def build_workbook_from_spec(
    data: dict,
    image_base_cli: Path | None = None,
) -> Workbook:
    """按范例几何重排新工作簿（避免对含图范本 openpyxl load/save 丢内嵌图）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"

    meta = data.get("meta") or {}
    scenes = data.get("scenes") or []
    props = data.get("props") or []
    characters = data.get("characters") or []

    base_meta = meta.get("配图根目录") or meta.get("image_base")
    base_path: Path | None = None
    if base_meta:
        base_path = Path(str(base_meta)).expanduser().resolve()
    if image_base_cli is not None:
        base_path = image_base_cli.resolve()

    max_side = float(meta.get("配图最大边长") or meta.get("image_max_side") or 320)

    widths = {
        "A": 5.775,
        "B": 7.183333,
        "C": 14.224999,
        "D": 35.891666,
        "E": 30.366667,
        "F": 18.0,
        "G": 2.5,  # 场景区与道具区之间的单列间隔
        "H": 9.566667,
        "I": 13.358334,
        "J": 27.658333,
        "K": 18.0,
        "L": 18.0,
        "M": 2.5,  # 道具区与角色区之间的单列间隔
        "N": 5.883333,
        "O": 10.0,
        "P": 18.0,
        "Q": 18.0,
        "R": 18.0,
        "S": 13.0,
        "T": 13.0,
        "U": 13.0,
        "V": 13.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bold = Font(bold=True)

    ws["B1"], ws["B2"], ws["B3"] = "编导", "脚本链接", "年级"
    ws["B1"].font = ws["B2"].font = ws["B3"].font = bold
    ws["F1"] = "定制互动色彩参考"
    ws["F1"].font = bold

    ws.merge_cells("C1:D1")
    ws.merge_cells("C2:D2")
    ws.merge_cells("C3:D3")
    ws["C1"] = meta.get("编导") or "—"
    ws["C2"] = meta.get("脚本链接") or "—"
    ws["C3"] = meta.get("年级") if meta.get("年级") not in (None, "") else "—"
    for addr in ("C1", "C2", "C3"):
        ws[addr].alignment = WRAP_TOP

    color_ref = meta.get("定制互动色彩参考")
    ws["F3"] = color_ref if color_ref not in (None, "") else "—"
    ws["F3"].alignment = WRAP_TOP

    ws["C8"], ws["K8"], ws["R8"] = "场景数量", "道具数量", "角色数量"
    for x in ("C8", "K8", "R8"):
        ws[x].font = bold
    # 表头数量仅统计「新增」：名称不以 -复用 结尾的行（与 JSON counts 无关，避免误填）
    n_scene, n_prop, n_char = _count_new_only_rows(scenes, props, characters)
    ws["D8"], ws["L8"], ws["S8"] = n_scene, n_prop, n_char

    scene_headers = list(SCENE_COLS.keys())
    for i, h in enumerate(scene_headers, start=1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = bold
        cell.alignment = HEADER_ALIGN
    prop_headers = list(PROP_COLS.keys())
    for i, h in enumerate(prop_headers, start=PROP_COLS["序号"]):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = bold
        cell.alignment = HEADER_ALIGN
    char_headers = list(CHAR_COLS.keys())
    for i, h in enumerate(char_headers, start=CHAR_COLS["序号"]):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = bold
        cell.alignment = HEADER_ALIGN

    # 不再写入「场景块说明」整段备注（旧 meta.场景块说明 / scene_block_note 已忽略），场景明细从第 11 行起与道具同排对齐
    scene_start = 11
    ws.cell(row=11, column=4, value=None)

    prop_row = 11
    scene_row = scene_start
    char_row = scene_start
    proto = scene_start

    image_tasks: list[tuple[int, dict[str, int], dict]] = []

    def _cell_val(d: dict, k: str) -> Any:
        v = d.get(k)
        if v is None or v == "":
            return None
        if isinstance(v, str):
            t = v.strip()
            if t in _PLACEHOLDER_STRIPS:
                return None
        return v

    def write_scene_row(row: int, d: dict) -> None:
        for k, col in SCENE_COLS.items():
            ws.cell(row=row, column=col, value=_cell_val(d, k))
        _style_data_row(ws, row, proto)
        if isinstance(d.get("配图") or d.get("images"), dict):
            image_tasks.append((row, SCENE_COLS, d))

    def write_prop_row(row: int, d: dict) -> None:
        for k, col in PROP_COLS.items():
            ws.cell(row=row, column=col, value=_cell_val(d, k))
        _style_data_row(ws, row, proto)
        if isinstance(d.get("配图") or d.get("images"), dict):
            image_tasks.append((row, PROP_COLS, d))

    def write_char_row(row: int, d: dict) -> None:
        for k, col in CHAR_COLS.items():
            ws.cell(row=row, column=col, value=_cell_val(d, k))
        _style_data_row(ws, row, proto)
        if isinstance(d.get("配图") or d.get("images"), dict):
            image_tasks.append((row, CHAR_COLS, d))

    for item in props:
        if isinstance(item, dict) and "__section__" in item:
            title = item["__section__"] or ""
            kc = PROP_COLS["道具名称"]
            _style_data_row(ws, prop_row, proto)
            cell = ws.cell(row=prop_row, column=kc, value=title)
            cell.font = Font(bold=True)
            cell.alignment = WRAP_TOP
            prop_row += 1
            continue
        if isinstance(item, dict):
            write_prop_row(prop_row, item)
            prop_row += 1

    for item in scenes:
        if isinstance(item, dict):
            write_scene_row(scene_row, item)
            scene_row += 1

    for item in characters:
        if isinstance(item, dict):
            write_char_row(char_row, item)
            char_row += 1

    for row, cmap, d in image_tasks:
        _embed_row_images(ws, row, cmap, d, base_path, max_side)

    return wb


def fill_template_inplace(template: Path, data: dict, output: Path) -> None:
    import shutil

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    print(
        f"已复制范本（保留内嵌图）至: {output.resolve()}\n"
        "提示：未自动改单元格内容；填数请用生成模式或 Excel 手工粘贴。",
        file=sys.stderr,
    )


def main() -> None:
    p = argparse.ArgumentParser(description="美术统筹表 JSON → xlsx（对齐范例版式，可嵌配图）")
    p.add_argument("input_json", type=Path, nargs="?", help="统筹数据 JSON 路径")
    p.add_argument("-o", "--output", type=Path, required=True, help="输出 .xlsx 路径")
    p.add_argument(
        "--template",
        type=Path,
        default=None,
        help="仅复制该 xlsx 到 -o（保留图片），不解析 JSON",
    )
    p.add_argument(
        "--copy-only",
        action="store_true",
        help="与 --template 联用：只复制文件",
    )
    p.add_argument(
        "--image-base",
        type=Path,
        default=None,
        help="配图相对路径解析目录（覆盖 meta.配图根目录）",
    )
    args = p.parse_args()

    if args.template is not None:
        if not args.template.is_file():
            print(f"找不到范本: {args.template}", file=sys.stderr)
            sys.exit(1)
        if args.copy_only or args.input_json is None:
            fill_template_inplace(args.template, {}, args.output)
            return

    if args.input_json is None:
        p.error("请提供 input_json，或使用 --template 与 --copy-only")

    raw = args.input_json.read_text(encoding="utf-8")
    data = json.loads(raw)
    wb = build_workbook_from_spec(data, image_base_cli=args.image_base)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"已写入: {args.output.resolve()}")


if __name__ == "__main__":
    main()
