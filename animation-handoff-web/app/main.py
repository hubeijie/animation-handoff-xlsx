"""
默认：浏览器直连 OpenAI 兼容 API 生成统筹 JSON，再 POST /api/build-xlsx 仅生成 .xlsx。
可选：勾选「服务端调用模型」时走 /api/convert（服务端持有 OPENAI_API_KEY）。

环境变量（仅服务端模式）：
  OPENAI_API_KEY       服务端代调模型时必填（除非 DRY_RUN=1）
  OPENAI_BASE_URL      默认 https://api.openai.com/v1
  OPENAI_MODEL         默认 gpt-4o
  GENERATOR_SCRIPT     默认优先仓库内 scripts/generate_art_coordination_xlsx.py，否则 ~/.cursor/skills/.../generate_art_coordination_xlsx.py
  DRY_RUN=1            不调用 API，返回固定最小 JSON 样例（仅测生成链路）
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path

import httpx
from fastapi import BackgroundTasks, Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

APP_ROOT = Path(__file__).resolve().parent.parent
STATIC_DIR = APP_ROOT / "static"
_BUNDLED_GENERATOR = APP_ROOT / "scripts" / "generate_art_coordination_xlsx.py"
_SKILL_GENERATOR = (
    Path.home()
    / ".cursor/skills/animation-handoff-xlsx/scripts/generate_art_coordination_xlsx.py"
)
DEFAULT_GENERATOR = (
    _BUNDLED_GENERATOR if _BUNDLED_GENERATOR.is_file() else _SKILL_GENERATOR
)

SYSTEM_PROMPT = """你是「小学动画课」美术统筹助手，须严格按 animation-handoff 规则从用户提供的脚本表格文字生成统筹数据。

【复用】仅当脚本某行出现 *复用场景-…* / *复用道具-…* / *复用人设-…* 时，对应「场景名称」「道具名称」「角色名称」须在归纳名后加半角后缀 -复用。不得凭「同上」「沿用」等口语推断复用。

【描述】场景/道具/角色描述列：连贯书面语给美术看；不要写【剧情】【画风】等标签前缀；须含画风建议、整体色调与冷暖；昼夜与主光要写清；不写剧情节奏概括句；知识点相关处可标 [待教研确认]；推断标 [推断]。描述中不要重复粘贴 *复用场景-* 等原文标记。

【输出】只输出一个 UTF-8 JSON 对象，不要 Markdown 代码围栏，不要注释。键名必须如下（可省略无值的图列）：

{
  "meta": {
    "编导": "string 或从表推断",
    "脚本链接": "— 或 URL",
    "年级": "string 或 —",
    "定制互动色彩参考": "— 或 string"
  },
  "scenes": [
    {
      "序号": 1,
      "背景文件名": null,
      "场景名称": "string",
      "场景描述": "string",
      "脚本截图": null,
      "参考图": null
    }
  ],
  "props": [
    { "__section__": "以下复用" },
    {
      "序号": 1,
      "道具名称": "string",
      "道具描述": "string",
      "脚本截图": null,
      "参考图": null
    }
  ],
  "characters": [
    {
      "序号": 1,
      "角色名称": "string",
      "角色描述": "string",
      "脚本截图": null,
      "参考图": null
    }
  ]
}

道具块里若脚本有「以下复用」等小节标题，用 {"__section__": "小节原文"} 插入 props 数组。无图则脚本截图、参考图、配图均省略或 null。

合并同一环境多镜为一行场景（除非脚本明确分子区域）。不新增脚本未出现的场景/道具；若必须推断则 [推断] 并简述依据。不改台词与【n】锚点文字（描述中不要复述大段台词）。"""


def xlsx_to_text(
    path: Path,
    max_rows_per_sheet: int = 500,
    sheet_name: str | None = None,
) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    want = (sheet_name or "").strip()
    sheets = wb.worksheets
    if want:
        sheets = [s for s in wb.worksheets if s.title.strip() == want]
        if not sheets:
            wb.close()
            names = "、".join(s.title for s in wb.worksheets)
            raise HTTPException(
                status_code=400,
                detail=f"未找到名为「{want}」的工作表。当前工作表：{names or '（无）'}",
            )
    for sheet in sheets:
        parts.append(f"## 工作表: {sheet.title}")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            row_count += 1
            if row_count > max_rows_per_sheet:
                parts.append(f"…（本表已截断，仅前 {max_rows_per_sheet} 行）")
                break
            line = "\t".join(
                "" if c is None else str(c).replace("\n", " ") for c in row
            )
            parts.append(line)
    wb.close()
    return "\n".join(parts)


def extract_json_object(raw: str) -> dict:
    s = raw.strip()
    fence = re.match(r"^```(?:json)?\s*([\s\S]*?)\s*```$", s)
    if fence:
        s = fence.group(1).strip()
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        start = s.find("{")
        end = s.rfind("}")
        if start >= 0 and end > start:
            return json.loads(s[start : end + 1])
        raise


def _merge_meta_from_form(data: dict, meta_hints: dict[str, str]) -> None:
    """用户填写的表头信息覆盖/补全 JSON meta（非空才写入）。"""
    meta = data.setdefault("meta", {})
    if not isinstance(meta, dict):
        data["meta"] = {}
        meta = data["meta"]
    for cn_key, val in meta_hints.items():
        v = (val or "").strip()
        if v:
            meta[cn_key] = v


async def llm_json(script_text: str, meta_hints: dict[str, str]) -> dict:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="未配置 OPENAI_API_KEY，无法生成统筹 JSON。",
        )
    base = os.environ.get("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")
    model = os.environ.get("OPENAI_MODEL", "gpt-4o")
    url = f"{base}/chat/completions"
    hint_lines = [
        f"{k}：{v}"
        for k, v in (
            ("编导", meta_hints.get("编导", "")),
            ("脚本链接", meta_hints.get("脚本链接", "")),
            ("年级", meta_hints.get("年级", "")),
            ("定制互动色彩参考", meta_hints.get("定制互动色彩参考", "")),
        )
        if (v or "").strip()
    ]
    hint_block = (
        "【表头信息】用户已填写下列项，请写入 JSON 的 meta 对应字段；若与脚本表格内信息冲突，以脚本为准。\n"
        + "\n".join(hint_lines)
        + "\n\n"
        if hint_lines
        else ""
    )
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": hint_block
                + "以下为脚本 Excel 导出的制表符分隔文本（含表头）。请生成完整 handoff JSON：\n\n"
                + script_text[:120_000],
            },
        ],
        "temperature": 0.2,
    }
    async with httpx.AsyncClient(timeout=300.0) as client:
        r = await client.post(
            url,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
    if r.status_code >= 400:
        raise HTTPException(
            status_code=502,
            detail=f"模型 API 错误 {r.status_code}: {r.text[:2000]}",
        )
    data = r.json()
    try:
        content = data["choices"][0]["message"]["content"]
    except (KeyError, IndexError) as e:
        raise HTTPException(status_code=502, detail=f"模型响应异常: {data!r}") from e
    return extract_json_object(content)


def dry_run_json(meta_hints: dict[str, str] | None = None) -> dict:
    meta = {
        "编导": "—",
        "脚本链接": "—",
        "年级": "—",
        "定制互动色彩参考": "—",
    }
    if meta_hints:
        for k, v in meta_hints.items():
            if (v or "").strip():
                meta[k] = v.strip()
    return {
        "meta": meta,
        "scenes": [
            {
                "序号": 1,
                "背景文件名": None,
                "场景名称": "示例教室",
                "场景描述": "儿童向平涂、柔和日光从左侧窗入，暖白墙面与浅木桌椅，前景留白给角色。[推断]",
            }
        ],
        "props": [],
        "characters": [
            {
                "序号": 1,
                "角色名称": "小主角",
                "角色描述": "简约线面风格，主色块清晰，与场景色调统一。[推断]",
            }
        ],
    }


def run_generator(data: dict, out_path: Path) -> None:
    gen = Path(
        os.environ.get("GENERATOR_SCRIPT", str(DEFAULT_GENERATOR))
    ).expanduser()
    if not gen.is_file():
        raise HTTPException(
            status_code=500,
            detail=f"找不到生成脚本: {gen}，请设置 GENERATOR_SCRIPT。",
        )
    tmp_json = out_path.with_suffix(".json")
    tmp_json.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    try:
        proc = subprocess.run(
            [
                "python3",
                str(gen),
                str(tmp_json),
                "-o",
                str(out_path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
    finally:
        tmp_json.unlink(missing_ok=True)
    if proc.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=f"生成 xlsx 失败:\n{proc.stderr or proc.stdout}",
        )


app = FastAPI(
    title="美术统筹表工坊",
    description="脚本 Excel 整理导出为美术统筹表 .xlsx",
)

if STATIC_DIR.is_dir():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/", response_class=HTMLResponse)
async def index() -> str:
    p = STATIC_DIR / "index.html"
    if not p.is_file():
        return "<p>缺少 static/index.html</p>"
    return p.read_text(encoding="utf-8")


@app.get("/api/handoff-system-prompt")
async def handoff_system_prompt() -> dict:
    """供浏览器直连大模型时拉取与后端一致的 system 提示词。"""
    return {"prompt": SYSTEM_PROMPT}


@app.post("/api/build-xlsx")
async def build_xlsx(
    background_tasks: BackgroundTasks,
    body: dict = Body(...),
) -> FileResponse:
    """仅根据统筹 JSON 生成 .xlsx（不调大模型）。浏览器直连模型后 POST 至此。"""
    data = body.get("data")
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="JSON 根对象须包含 data 字段（统筹对象）")
    stem_raw = body.get("filename_stem")
    stem = str(stem_raw).strip() if stem_raw else "脚本"
    dl_name = f"{_safe_download_stem(stem + '.xlsx')}_美术统筹表.xlsx"

    tmp = Path(tempfile.gettempdir()) / f"ahw-{uuid.uuid4().hex}"
    tmp.mkdir(parents=True, exist_ok=True)
    out_path = tmp / "out.xlsx"

    def _rm_tmp() -> None:
        shutil.rmtree(tmp, ignore_errors=True)

    try:
        run_generator(data, out_path)
        background_tasks.add_task(_rm_tmp)
        return FileResponse(
            path=str(out_path),
            filename=dl_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        _rm_tmp()
        raise


def _safe_download_stem(name: str | None) -> str:
    base = Path(name or "脚本").stem.strip() or "脚本"
    base = re.sub(r'[\s<>"|:*?\\/]+', "_", base)
    return base[:72] if len(base) > 72 else base


@app.post("/api/convert")
async def convert(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    director: str = Form(""),
    script_url: str = Form(""),
    grade: str = Form(""),
    color_ref: str = Form(""),
    sheet_name: str = Form(""),
    max_rows: str = Form("500"),
) -> FileResponse:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")

    try:
        nrows = max(50, min(int((max_rows or "500").strip()), 2000))
    except ValueError:
        raise HTTPException(status_code=400, detail="「每个工作表最多读取行数」须为 50～2000 的整数")
    meta_hints = {
        "编导": director or "",
        "脚本链接": script_url or "",
        "年级": grade or "",
        "定制互动色彩参考": color_ref or "",
    }
    sheet = sheet_name.strip() or None

    tmp = Path(tempfile.gettempdir()) / f"ahw-{uuid.uuid4().hex}"
    tmp.mkdir(parents=True, exist_ok=True)
    in_path = tmp / "input.xlsx"
    out_path = tmp / "out.xlsx"
    dl_name = f"{_safe_download_stem(file.filename)}_美术统筹表.xlsx"

    def _rm_tmp() -> None:
        shutil.rmtree(tmp, ignore_errors=True)

    try:
        with in_path.open("wb") as f:
            shutil.copyfileobj(file.file, f)

        script_text = xlsx_to_text(in_path, max_rows_per_sheet=nrows, sheet_name=sheet)
        if os.environ.get("DRY_RUN") == "1":
            data = dry_run_json(meta_hints)
        else:
            data = await llm_json(script_text, meta_hints)
            _merge_meta_from_form(data, meta_hints)

        run_generator(data, out_path)
        background_tasks.add_task(_rm_tmp)
        return FileResponse(
            path=str(out_path),
            filename=dl_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        _rm_tmp()
        raise
    finally:
        await file.close()
